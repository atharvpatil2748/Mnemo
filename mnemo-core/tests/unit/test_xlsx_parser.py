"""Behavioral tests for the generic XLSX parser."""

import hashlib
import io

import pytest
from mnemo.interfaces.errors import ContractValidationError
from mnemo.interfaces.parser_models import RawHeadingBlock, RawTableBlock
from mnemo.interfaces.types import FileMetadata
from mnemo.models import FrozenMetadata
from mnemo.parsers import XLSXParser
from openpyxl import Workbook  # type: ignore[import-untyped]


def _metadata(data: bytes) -> FileMetadata:
    return FileMetadata(
        content_hash=hashlib.sha256(data).hexdigest(),
        size_bytes=len(data),
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        modified_at=None,
        metadata=FrozenMetadata(),
    )


def _workbook_bytes() -> bytes:
    workbook = Workbook()
    first = workbook.active
    first.title = "Schedule"
    first.append(("Name", "Day", "Active", "Formula"))
    first.append(("A", "Monday", True, "=1+1"))
    first.append((None, None, None, None))
    first.append(("B", "Tuesday", False, 3.5))
    second = workbook.create_sheet("Empty")
    second["A1"] = None
    stream = io.BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def test_xlsx_parser_preserves_sheet_rows_values_and_formulas() -> None:
    data = _workbook_bytes()
    result = XLSXParser().parse(data, "schedule.xlsx", _metadata(data))

    assert result.metadata.title == "schedule"
    assert result.metadata.metadata["sheet_count"] == 1
    assert result.metadata.metadata["sheet_names"] == ("Schedule",)
    assert isinstance(result.blocks[0], RawHeadingBlock)
    assert result.blocks[0].text == "Schedule"
    table = result.blocks[1]
    assert isinstance(table, RawTableBlock)
    assert table.rows == (
        ("Name", "Day", "Active", "Formula"),
        ("A", "Monday", "TRUE", "=1+1"),
        ("B", "Tuesday", "FALSE", "3.5"),
    )
    assert table.header_row_count == 1


def test_xlsx_parser_partitions_rows_without_loss() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(("ID", "Value"))
    for index in range(120):
        sheet.append((index, f"row-{index}"))
    stream = io.BytesIO()
    workbook.save(stream)
    workbook.close()
    data = stream.getvalue()

    result = XLSXParser().parse(data, "large.xlsx", _metadata(data))
    tables = tuple(block for block in result.blocks if isinstance(block, RawTableBlock))
    reconstructed = [*tables[0].rows]
    for table in tables[1:]:
        assert table.rows[0] == ("ID", "Value")
        reconstructed.extend(table.rows[1:])
    assert len(tables) == 3
    assert len(reconstructed) == 121
    assert reconstructed[-1] == ("119", "row-119")


@pytest.mark.parametrize("data", [b"", b"not-an-office-archive"])
def test_xlsx_parser_rejects_empty_or_malformed_input(data: bytes) -> None:
    with pytest.raises(ContractValidationError):
        XLSXParser().parse(data, "bad.xlsx", _metadata(data))
