"""Built-in parser for Office Open XML workbooks (.xlsx)."""

from __future__ import annotations

import io
from collections.abc import Iterable
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook  # type: ignore[import-untyped]

from mnemo.interfaces.errors import ContractValidationError
from mnemo.interfaces.parser import ParserInterfaceV1
from mnemo.interfaces.parser_models import ParseResult, RawBlock, RawHeadingBlock, RawTableBlock
from mnemo.interfaces.types import FileMetadata, ParserCapabilities
from mnemo.models import DocType, DocumentMetadata
from mnemo.models._shared import FrozenMetadata

_MAX_ROWS_PER_BLOCK = 50


def _cell_text(value: Any) -> str:
    """Return one deterministic, searchable cell representation."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)


def _normalized_rows(rows: Iterable[tuple[Any, ...]]) -> tuple[tuple[str, ...], ...]:
    """Discard only wholly empty rows and pad remaining rows to a stable width."""
    materialized = [tuple(_cell_text(cell) for cell in row) for row in rows]
    meaningful = [row for row in materialized if any(cell != "" for cell in row)]
    if not meaningful:
        return ()
    width = max(len(row) for row in meaningful)
    return tuple(row + ("",) * (width - len(row)) for row in meaningful)


def _table_partitions(rows: tuple[tuple[str, ...], ...]) -> tuple[tuple[tuple[str, ...], ...], ...]:
    """Partition large sheets without splitting or dropping logical rows."""
    if len(rows) <= _MAX_ROWS_PER_BLOCK:
        return (rows,)
    header = rows[0]
    data = rows[1:]
    capacity = _MAX_ROWS_PER_BLOCK - 1
    return tuple(
        (header, *data[start : start + capacity]) for start in range(0, len(data), capacity)
    )


class XLSXParser(ParserInterfaceV1):
    """Parse workbook sheets into titled, row-preserving table blocks."""

    @property
    def supported_formats(self) -> tuple[str, ...]:
        return (
            ".xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def capabilities(self) -> ParserCapabilities:
        return ParserCapabilities(
            supported_formats=self.supported_formats,
            supports_images=False,
            supports_tables=True,
            supports_math=False,
            supports_ocr=False,
        )

    def parse(self, data: bytes, filename: str, metadata: FileMetadata) -> ParseResult:
        """Parse exact workbook bytes without evaluating formulas or following links."""
        if not data:
            raise ContractValidationError(f"Cannot parse empty XLSX: {filename}")

        try:
            workbook = load_workbook(
                io.BytesIO(data),
                read_only=True,
                data_only=False,
                keep_links=False,
            )
        except Exception as error:
            raise ContractValidationError(f"Failed to open XLSX workbook: {error}") from error

        blocks: list[RawBlock] = []
        sheet_names: list[str] = []
        ordinal = 0
        try:
            for worksheet in workbook.worksheets:
                rows = _normalized_rows(worksheet.iter_rows(values_only=True))
                if not rows:
                    continue
                sheet_names.append(worksheet.title)
                blocks.append(
                    RawHeadingBlock(
                        ordinal=ordinal,
                        text=worksheet.title,
                        level=1,
                    )
                )
                ordinal += 1
                for partition in _table_partitions(rows):
                    blocks.append(
                        RawTableBlock(
                            ordinal=ordinal,
                            rows=partition,
                            header_row_count=1 if len(partition) > 1 else 0,
                        )
                    )
                    ordinal += 1
        finally:
            workbook.close()

        title = Path(filename).stem if filename else "Untitled"
        document_metadata = DocumentMetadata(
            content_hash=metadata.content_hash,
            title=title,
            metadata=FrozenMetadata(
                {
                    **dict(metadata.metadata),
                    "sheet_count": len(sheet_names),
                    "sheet_names": tuple(sheet_names),
                }
            ),
        )
        return ParseResult(
            blocks=tuple(blocks),
            extracted_assets=(),
            metadata=document_metadata,
            language="en",
            doc_type=DocType.GENERIC,
        )
